"""Document ingestion: parse → chunk → embed → store in Qdrant.

Supports three embedding providers controlled by EMBEDDING_PROVIDER in .env:

  openai   (default) — text-embedding-3-small  → requires OPENAI_API_KEY
                        text-embedding-3-large, text-embedding-ada-002
  gemini             — models/text-embedding-004 → requires GEMINI_API_KEY
  local              — sentence-transformers all-MiniLM-L6-v2 (384-dim)
                        runs entirely offline; install:
                        pip install sentence-transformers

Vector dimensions per provider:
  openai / text-embedding-3-small  → 1536
  openai / text-embedding-3-large  → 3072
  gemini / text-embedding-004      → 768
  local  / all-MiniLM-L6-v2        → 384
"""
from __future__ import annotations

import hashlib
import re
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional

from app.core.config import settings
from app.core.exceptions import DocumentIngestionError

import logging
log = logging.getLogger(__name__)


@dataclass
class EmbeddingConfig:
    """Per-project embedding configuration.  Falls back to global settings if any field is empty."""
    provider: str = ""
    model: str = ""
    api_key: str = ""

    def resolved_provider(self) -> str:
        return self.provider or settings.EMBEDDING_PROVIDER

    def resolved_model(self) -> str:
        return self.model or settings.EMBEDDING_MODEL

    def resolved_key(self) -> str:
        return self.api_key  # callers fall back to settings key when empty


# ─────────────────────────────── Qdrant helpers ──────────────────────────────

def _qdrant():
    from qdrant_client import QdrantClient
    if settings.QDRANT_URL:
        kwargs = {"url": settings.QDRANT_URL}
        if settings.QDRANT_API_KEY:
            kwargs["api_key"] = settings.QDRANT_API_KEY
        return QdrantClient()
    import os
    os.makedirs(settings.QDRANT_PATH, exist_ok=True)
    return QdrantClient(path=settings.QDRANT_PATH)


def _collection_name(project_id: str) -> str:
    return f"project_{project_id.replace('-', '_')}"


def _vector_size(cfg: Optional[EmbeddingConfig] = None) -> int:
    """Return the vector dimension for the configured embedding provider/model."""
    p = (cfg.resolved_provider() if cfg else settings.EMBEDDING_PROVIDER).lower()
    m = (cfg.resolved_model() if cfg else settings.EMBEDDING_MODEL).lower()
    if p == "local":
        return 384
    if p == "gemini":
        if "embedding-001" in m or "embedding-exp" in m:
            return 3072
        return 768
    # OpenAI
    if "large" in m:
        return 3072
    return 1536  # small / ada-002


def _ensure_collection(client, name: str, cfg: Optional[EmbeddingConfig] = None) -> None:
    from qdrant_client.models import Distance, VectorParams
    expected = _vector_size(cfg)
    existing = {c.name for c in client.get_collections().collections}
    if name in existing:
        info = client.get_collection(name)
        actual = info.config.params.vectors.size
        if actual != expected:
            log.warning(
                "Qdrant collection %s has vector size %d but current model needs %d — "
                "recreating collection (existing indexed data will be lost).",
                name, actual, expected,
            )
            client.delete_collection(name)
        else:
            return  # collection exists and size matches
    client.create_collection(
        collection_name=name,
        vectors_config=VectorParams(size=expected, distance=Distance.COSINE),
    )


# ─────────────────────────────── Embedding ───────────────────────────────────

def _embed_openai(texts: List[str], model: str = "", api_key: str = "") -> List[List[float]]:
    key = api_key or settings.OPENAI_API_KEY
    if not key:
        raise DocumentIngestionError(
            "EMBEDDING_PROVIDER=openai but no API key found.\n"
            "Set OPENAI_API_KEY in .env or add a project-level key in Settings."
        )
    resolved_model = model or settings.EMBEDDING_MODEL
    import openai
    client = openai.OpenAI(api_key=key)
    resp = client.embeddings.create(model=resolved_model, input=texts)
    return [item.embedding for item in resp.data]


def _embed_gemini(texts: List[str], model: str = "", api_key: str = "") -> List[List[float]]:
    key = api_key or settings.GEMINI_API_KEY
    if not key:
        raise DocumentIngestionError(
            "EMBEDDING_PROVIDER=gemini but no API key found.\n"
            "Set GEMINI_API_KEY in .env or add a project-level key in Settings."
        )
    import google.generativeai as genai
    genai.configure(api_key=key)
    resolved_model = model or settings.EMBEDDING_MODEL
    if resolved_model == "text-embedding-3-small":
        resolved_model = "models/text-embedding-004"
    results = []
    for text in texts:
        r = genai.embed_content(model=resolved_model, content=text, task_type="retrieval_document")
        results.append(r["embedding"])
    return results


_local_model = None  # module-level singleton to avoid reloading

# def _embed_local(texts: List[str]) -> List[List[float]]:
#     global _local_model
#     try:
#         from sentence_transformers import SentenceTransformer
#     except ImportError:
#         raise DocumentIngestionError(
#             "EMBEDDING_PROVIDER=local requires sentence-transformers.\n"
#             "Run: pip install sentence-transformers"
#         )
#     if _local_model is None:
#         model_name = settings.EMBEDDING_MODEL \
#             if settings.EMBEDDING_MODEL != "text-embedding-3-small" \
#             else "all-MiniLM-L6-v2"
#         log.info(f"loading local embedding model: {model_name}")
#         _local_model = SentenceTransformer(model_name)
#     vecs = _local_model.encode(texts, show_progress_bar=False)
#     return [v.tolist() for v in vecs]


def _embed(texts: List[str], cfg: Optional[EmbeddingConfig] = None) -> List[List[float]]:
    """Dispatch to the configured embedding provider."""
    p = (cfg.resolved_provider() if cfg else settings.EMBEDDING_PROVIDER).lower()
    m = cfg.resolved_model() if cfg else settings.EMBEDDING_MODEL
    k = cfg.resolved_key() if cfg else ""
    if p == "gemini":
        return _embed_gemini(texts, model=m, api_key=k)
    # if p == "local":
    #     return _embed_local(texts)
    return _embed_openai(texts, model=m, api_key=k)  # default


def _check_embedding_config(cfg: Optional[EmbeddingConfig] = None) -> None:
    """Raise a clear error early if the embedding provider isn't configured."""
    p = (cfg.resolved_provider() if cfg else settings.EMBEDDING_PROVIDER).lower()
    k = cfg.resolved_key() if cfg else ""
    if p == "openai" and not (k or settings.OPENAI_API_KEY):
        raise DocumentIngestionError(
            "Cannot index document: no OpenAI API key found.\n\n"
            "Quick fix — choose one option:\n"
            "  A) Add a project-level API key in Settings → AI / Embedding Settings\n"
            "  B) Add OPENAI_API_KEY to .env\n"
            "  C) Switch EMBEDDING_PROVIDER to 'gemini' or 'local'"
        )
    if p == "gemini" and not (k or settings.GEMINI_API_KEY):
        raise DocumentIngestionError(
            "EMBEDDING_PROVIDER=gemini but no Gemini API key found.\n"
            "Add a project-level key or set GEMINI_API_KEY in .env."
        )


# ─────────────────────────────── Text extraction ─────────────────────────────

def _extract_text(file_path: str) -> str:
    path = Path(file_path)
    ext  = path.suffix.lower().lstrip(".")

    if ext == "pdf":
        from pypdf import PdfReader
        reader = PdfReader(file_path)
        return "\n\n".join(page.extract_text() or "" for page in reader.pages)

    if ext == "docx":
        from docx import Document
        doc = Document(file_path)
        return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())

    if ext == "xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        lines = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join(str(c) for c in row if c is not None)
                if row_text.strip():
                    lines.append(row_text)
        return "\n".join(lines)

    return path.read_text(encoding="utf-8", errors="replace")


# ─────────────────────────────── Chunking ────────────────────────────────────

def _chunk_text(text: str, size: int = 0, overlap: int = 0) -> List[str]:
    size    = size    or settings.CHUNK_SIZE
    overlap = overlap or settings.CHUNK_OVERLAP

    paragraphs = re.split(r"\n\s*\n", text)
    chunks: List[str] = []
    buf = ""

    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        if len(buf) + len(para) + 2 <= size:
            buf = (buf + "\n\n" + para).strip()
        else:
            if buf:
                chunks.append(buf)
            while len(para) > size:
                chunks.append(para[:size])
                para = para[size - overlap:]
            buf = para

    if buf:
        chunks.append(buf)

    return [c for c in chunks if c.strip()]


# ─────────────────────────────── Public API ──────────────────────────────────

def ingest_document(
    file_path: str,
    project_id: str,
    document_id: str,
    cfg: Optional[EmbeddingConfig] = None,
) -> int:
    """Parse, chunk, embed, and upsert document into Qdrant. Returns chunk count."""
    filename = Path(file_path).name
    provider = cfg.resolved_provider() if cfg else settings.EMBEDDING_PROVIDER
    log.info(f"ingest_start file={filename} provider={provider}")

    _check_embedding_config(cfg)

    try:
        text = _extract_text(file_path)
    except Exception as exc:
        raise DocumentIngestionError(f"Parsing failed for {filename}: {exc}") from exc

    if not text.strip():
        raise DocumentIngestionError(f"No text extracted from {filename}")

    chunks = _chunk_text(text)
    if not chunks:
        raise DocumentIngestionError(f"Zero chunks produced for {filename}")

    log.info(f"ingest_chunked file={filename} chunks={len(chunks)}")

    try:
        BATCH = 200 if provider.lower() == "openai" else 32
        all_embeddings: List[List[float]] = []
        for i in range(0, len(chunks), BATCH):
            all_embeddings.extend(_embed(chunks[i : i + BATCH], cfg))
    except DocumentIngestionError:
        raise
    except Exception as exc:
        raise DocumentIngestionError(f"Embedding failed for {filename}: {exc}") from exc

    try:
        from qdrant_client.models import PointStruct
        client = _qdrant()
        col    = _collection_name(project_id)
        _ensure_collection(client, col, cfg)

        points = [
            PointStruct(
                id=str(uuid.uuid5(uuid.NAMESPACE_URL, f"{document_id}_{i}")),
                vector=all_embeddings[i],
                payload={
                    "document_id": document_id,
                    "chunk_index": i,
                    "text": chunks[i],
                    "source": filename,
                },
            )
            for i in range(len(chunks))
        ]
        for i in range(0, len(points), 100):
            client.upsert(collection_name=col, points=points[i : i + 100])
    except Exception as exc:
        raise DocumentIngestionError(f"Qdrant upsert failed for {filename}: {exc}") from exc

    log.info(f"ingest_complete file={filename} chunks={len(chunks)}")
    return len(chunks)


def retrieve_context(
    query: str,
    project_id: str,
    n: int = 0,
    cfg: Optional[EmbeddingConfig] = None,
) -> List[str]:
    """Return top-n text chunks most relevant to query. Never raises."""
    k = n or settings.RAG_TOP_K
    p = (cfg.resolved_provider() if cfg else settings.EMBEDDING_PROVIDER).lower()
    ek = cfg.resolved_key() if cfg else ""
    if p == "openai" and not (ek or settings.OPENAI_API_KEY):
        return []
    if p == "gemini" and not (ek or settings.GEMINI_API_KEY):
        return []
    try:
        client = _qdrant()
        col    = _collection_name(project_id)
        existing = {c.name for c in client.get_collections().collections}
        if col not in existing:
            return []
        q_vec = _embed([query], cfg)[0]
        hits  = client.search(collection_name=col, query_vector=q_vec, limit=k)
        return [h.payload.get("text", "") for h in hits if h.payload]
    except Exception as exc:
        log.warning(f"retrieve_context failed project={project_id} error={exc}")
        return []


def delete_document_vectors(project_id: str, document_id: str) -> int:
    """Delete all Qdrant vectors for a document."""
    try:
        from qdrant_client.models import Filter, FieldCondition, MatchValue
        client = _qdrant()
        col    = _collection_name(project_id)
        existing = {c.name for c in client.get_collections().collections}
        if col not in existing:
            return 0
        client.delete(
            collection_name=col,
            points_selector=Filter(
                must=[FieldCondition(key="document_id", match=MatchValue(value=document_id))]
            ),
        )
        log.info(f"vectors_deleted document={document_id}")
        return 1
    except Exception as exc:
        log.warning(f"delete_vectors failed document={document_id} error={exc}")
        return 0


def get_index_stats(project_id: str, cfg: Optional[EmbeddingConfig] = None) -> dict:
    """Return Qdrant collection stats for the project."""
    provider = cfg.resolved_provider() if cfg else settings.EMBEDDING_PROVIDER
    model    = cfg.resolved_model()    if cfg else settings.EMBEDDING_MODEL
    if provider == "gemini" and model == "text-embedding-3-small":
        model = "text-embedding-004"
    elif provider == "local" and model == "text-embedding-3-small":
        model = "all-MiniLM-L6-v2"

    base = {
        "embedding_model": f"{provider}/{model}",
        "vector_store": "Qdrant (local)" if not settings.QDRANT_URL else f"Qdrant @ {settings.QDRANT_URL}",
    }
    try:
        client = _qdrant()
        col    = _collection_name(project_id)
        existing = {c.name for c in client.get_collections().collections}
        if col not in existing:
            return {"total_vectors": 0, **base}
        info = client.get_collection(col)
        return {"total_vectors": info.vectors_count or 0, **base}
    except Exception:
        return {"total_vectors": 0, **base}
