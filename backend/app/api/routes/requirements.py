"""Fetch GitLab issues and manage requirements."""
from __future__ import annotations

import io
from typing import Any, Dict

from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_project_repo, get_requirement_repo, get_testcase_repo
from app.core.exceptions import GitLabError, NotFoundError
from app.db.repository import ProjectRepository, RequirementRepository, TestCaseRepository
from app.db.session import get_db
from app.schemas.schemas import RequirementFetch, RequirementRead


def _build_excel_bytes(requirement_title: str, test_cases: list) -> bytes:
    """Generate an Excel workbook from test cases and return raw bytes.

    Manual test cases are expanded into structured rows (Preconditions, Test Data,
    Step #, Action, Expected Result).  BDD test cases go on a separate sheet.
    """
    import json as _json
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.worksheet import Worksheet

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    def _style_header(ws: Worksheet, headers: list, widths: list) -> None:
        from openpyxl.utils import get_column_letter
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w

    manual_tcs = [tc for tc in test_cases if tc.format == "MANUAL"]
    bdd_tcs    = [tc for tc in test_cases if tc.format == "BDD"]

    wb = Workbook()
    first_sheet_used = False

    if manual_tcs:
        ws_m: Worksheet = wb.worksheets[0]
        ws_m.title = "Manual Test Cases"
        first_sheet_used = True
        headers_m = ["#", "Title", "Priority", "Type", "Tags",
                     "Preconditions", "Test Data", "Step #", "Action", "Expected Result"]
        widths_m  = [5, 38, 10, 14, 20, 32, 32, 7, 45, 45]
        _style_header(ws_m, headers_m, widths_m)
        for idx, tc in enumerate(manual_tcs, 1):
            try:
                data = _json.loads(tc.content)
            except Exception:
                data = {}
            preconditions = data.get("preconditions") or ""
            test_data     = data.get("test_data") or ""
            steps = data.get("steps") or []
            if not steps:
                steps = [{"action": "", "expected": ""}]
            for si, step in enumerate(steps, 1):
                ws_m.append([
                    idx, tc.title, tc.priority, tc.scenario_type or "", tc.tags or "",
                    preconditions, test_data, si,
                    step.get("action", ""), step.get("expected", ""),
                ])

    if bdd_tcs:
        if not first_sheet_used:
            ws_b: Worksheet = wb.worksheets[0]
            first_sheet_used = True
        else:
            ws_b = wb.create_sheet()
        ws_b.title = "BDD Scenarios"
        headers_b = ["#", "Title", "Priority", "Type", "Tags", "Content"]
        widths_b  = [5, 38, 10, 14, 20, 80]
        _style_header(ws_b, headers_b, widths_b)
        for idx, tc in enumerate(bdd_tcs, 1):
            ws_b.append([idx, tc.title, tc.priority,
                         tc.scenario_type or "", tc.tags or "", tc.content])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

router = APIRouter()


@router.post("/fetch", response_model=RequirementRead, status_code=201)
async def fetch_requirement(
    payload: RequirementFetch,
    req_repo: RequirementRepository = Depends(get_requirement_repo),
    proj_repo: ProjectRepository = Depends(get_project_repo),
    db: AsyncSession = Depends(get_db),
):
    """Fetch one GitLab issue. Idempotent — returns existing record if already fetched."""
    try:
        project = await proj_repo.get_or_raise(payload.project_id)
    except NotFoundError as exc:
        raise HTTPException(404, str(exc))

    existing = await req_repo.find_by_issue_id(payload.project_id, payload.gitlab_issue_id)
    if existing:
        return existing

    from app.services.gitlab_service import GitLabService
    try:
        data = GitLabService(project).fetch_issue(payload.gitlab_issue_id)
    except GitLabError as exc:
        raise HTTPException(422, str(exc))

    req = await req_repo.create(project_id=payload.project_id, **data)
    await db.commit()
    await db.refresh(req)
    return req


@router.post("/bulk-fetch/{project_id}", response_model=list[RequirementRead], status_code=201)
async def bulk_fetch(
    project_id: str,
    req_repo: RequirementRepository = Depends(get_requirement_repo),
    proj_repo: ProjectRepository = Depends(get_project_repo),
    db: AsyncSession = Depends(get_db),
):
    """Import all GitLab issues matching the project's label/state filters."""
    try:
        project = await proj_repo.get_or_raise(project_id)
    except NotFoundError as exc:
        raise HTTPException(404, str(exc))

    from app.services.gitlab_service import GitLabService
    try:
        issues = GitLabService(project).list_issues(
            labels=project.label_include,
            state=project.issue_state,
            max_results=project.max_issues,
        )
    except GitLabError as exc:
        raise HTTPException(422, str(exc))

    created = []
    for data in issues:
        if await req_repo.find_by_issue_url(project_id, data["gitlab_issue_url"]):
            continue
        req = await req_repo.create(project_id=project_id, **data)
        created.append(req)

    await db.commit()
    return created


@router.get("/project/{project_id}", response_model=list[RequirementRead])
async def list_requirements(
    project_id: str,
    proj_repo: ProjectRepository = Depends(get_project_repo),
    req_repo: RequirementRepository = Depends(get_requirement_repo),
):
    try:
        await proj_repo.get_or_raise(project_id)
    except NotFoundError as exc:
        raise HTTPException(404, str(exc))
    return await req_repo.list_by_project(project_id)


@router.get("/{requirement_id}", response_model=RequirementRead)
async def get_requirement(
    requirement_id: str,
    req_repo: RequirementRepository = Depends(get_requirement_repo),
):
    try:
        return await req_repo.get_or_raise(requirement_id)
    except NotFoundError as exc:
        raise HTTPException(404, str(exc))


@router.delete("/{requirement_id}", status_code=204)
async def delete_requirement(
    requirement_id: str,
    req_repo: RequirementRepository = Depends(get_requirement_repo),
    db: AsyncSession = Depends(get_db),
):
    try:
        req = await req_repo.get_or_raise(requirement_id)
    except NotFoundError as exc:
        raise HTTPException(404, str(exc))
    await req_repo.delete(req)
    await db.commit()


@router.post("/{requirement_id}/push-to-gitlab")
async def push_to_gitlab(
    requirement_id: str,
    req_repo: RequirementRepository = Depends(get_requirement_repo),
    proj_repo: ProjectRepository = Depends(get_project_repo),
    tc_repo: TestCaseRepository = Depends(get_testcase_repo),
) -> Dict[str, Any]:
    """Post a markdown test-case summary + Excel attachment to the linked GitLab issue."""
    try:
        req = await req_repo.get_or_raise(requirement_id)
    except NotFoundError as exc:
        raise HTTPException(404, str(exc))

    if not req.gitlab_issue_url:
        raise HTTPException(400, "Requirement has no linked GitLab issue URL")

    try:
        project = await proj_repo.get_or_raise(req.project_id)
    except NotFoundError as exc:
        raise HTTPException(404, str(exc))

    test_cases = list(await tc_repo.list_by_requirement(requirement_id))
    if not test_cases:
        raise HTTPException(400, "No test cases to upload — generate some first")

    excel_bytes = _build_excel_bytes(req.title, test_cases)
    excel_filename = f"testcases_issue_{req.gitlab_issue_id or 'export'}.xlsx"

    from app.services.gitlab_service import GitLabService
    try:
        result = GitLabService(project).post_test_cases_to_issue(
            issue_url=req.gitlab_issue_url,
            requirement_title=req.title,
            test_cases=test_cases,
            excel_bytes=excel_bytes,
            excel_filename=excel_filename,
        )
    except GitLabError as exc:
        raise HTTPException(422, str(exc))

    return result
